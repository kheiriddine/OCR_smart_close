import json
import csv
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime
import subprocess
import together
import os

def convert_date_to_french_format(dt):
    """Convert datetime object or ISO string to French date format (dd/mm/YYYY)"""
    if isinstance(dt, datetime):
        return dt.strftime('%d/%m/%Y')
    if isinstance(dt, str):
        try:
            dt_obj = datetime.fromisoformat(dt.replace('Z', ''))
            return dt_obj.strftime('%d/%m/%Y')
        except ValueError:
            return dt
    return dt

def clean_value(value, field_type):
    """Clean and convert values based on field type"""
    if value is None or value == '':
        return 0 if field_type in ('int', 'float') else None

    if field_type == 'int':
        if isinstance(value, str) and value.startswith('="') and value.endswith('"'):
            value = value[2:-1]
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return 0

    if field_type == 'float':
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0

    if field_type == 'date':
        return convert_date_to_french_format(value)

    return str(value).strip() if value else None

def convert_to_json(input_path, output_path=None):
    path = Path(input_path)
    output_path = output_path if output_path else path.with_suffix('.json')
    ext = path.suffix.lower()

    if ext == '.csv':
        with open(input_path, encoding='utf-8') as f:
            data = list(csv.DictReader(f))
            for row in data:
                for key in row:
                    if key in ['N° pièce', 'N° compte général']:
                        row[key] = clean_value(row[key], 'int')
                    elif key in ['Débit', 'Crédit', 'P']:
                        row[key] = clean_value(row[key], 'float')
                    elif 'Date' in key:
                        row[key] = clean_value(row[key], 'date')
                    else:
                        row[key] = clean_value(row[key], 'str')

    elif ext in ('.xlsx', '.xls'):
        wb = load_workbook(filename=input_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        data = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = {}
            for header, value in zip(headers, row):
                if header in ['N° pièce', 'N° compte général']:
                    row_data[header] = clean_value(value, 'int')
                elif header in ['Débit', 'Crédit', 'P']:
                    row_data[header] = clean_value(value, 'float')
                elif 'Date' in str(header):
                    row_data[header] = clean_value(value, 'date')
                else:
                    row_data[header] = clean_value(value, 'str')
            # Remove null entries
            row_data = {k: v for k, v in row_data.items() if v is not None}
            data.append(row_data)

    elif ext == '.xml':
        def parse(element):
            d = element.attrib.copy()
            for child in element:
                child_data = parse(child)
                if child.tag in d:
                    if isinstance(d[child.tag], list):
                        d[child.tag].append(child_data)
                    else:
                        d[child.tag] = [d[child.tag], child_data]
                else:
                    d[child.tag] = child_data
            return d if d else element.text.strip() if element.text else None

        data = parse(ET.parse(input_path).getroot())

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=4, ensure_ascii=False)

def run_llama_ocr(pdf_path, output_file, temp_images_dir, ocr_api_key):
    try:
        # Modified to process a single file path instead of directory
        subprocess.run(["node", "llama_ocr_pdfsandimages.js", pdf_path, output_file, temp_images_dir, ocr_api_key], check=True)
        return output_file
    except subprocess.CalledProcessError as e:
        print(f"An error occurred: {e}")
        return None

def extract_info_from_ocr(ocr_output_file, together_api_key, original_filename):
    if not ocr_output_file:
        print("No OCR output file provided.")
        return

    with open(ocr_output_file, "r", encoding="utf-8") as f:
        content = f.read()

    prompt = f"""
    Tu es un agent intelligent qui extrait des informations financières structurées depuis des documents OCR.

    Voici le texte extrait :

    {content}

    Ta tâche est de retourner un objet JSON avec trois champs :
    1. `Nom Societe` : le nom de la société émettrice (trouvé dans l'en-tête/logo, généralement en haut de chaque page)
    2. `payment_info` : les infos de paiement, comme dans l'exemple ci-dessous.
    3. `table` : une table combinée extraite à partir de toutes les tables visibles dans le texte , au format JSON (type pandas DataFrame).

    Exemple de format :

    {{  "Nom Societe":"...", // nom de la société émettrice (pas le fournisseur)
        "payment_info": {{
            "Nom Fournisseur": "...",  // Nom exact du fournisseur bénéficiaire
            "Adresse Fournisseur": "...",
            ...,
            "Total":...
        }},
        "table": [
            {{ "Colonne1": "valeur1", "Colonne2": "valeur2", ... }},
            ...
        ]
    }}
    Règles de formatage STRICTES :
    1. Nombres/montants :
       - Toujours en format float (point comme séparateur décimal)
       - Supprimer les séparateurs de milliers
       - Exemple : "378 885,00 €" → 378885.00

    2. Dates :
       - Toujours en format ISO français (JJ/MM/AAAA)

    3. Numéros de référence :
       - Garder exactement comme dans le texte original
       - Ne pas modifier les tirets ou espaces
       - Exemple : "XXX-XXX-000-TOG-2867" → "XXX-XXX-000-TOG-2867"

    4. Adresses :
       - Garder la casse originale
       - Ne pas modifier la ponctuation

    Retourne uniquement le JSON & concatenater les deux tables des differents page existe dans le fichier .md .
    """

    client = together.Together(api_key=together_api_key)
    response = client.chat.completions.create(
        model="meta-llama/Llama-3.3-70b-Instruct-Turbo-Free",
        messages=[{"role": "user", "content": prompt}],
        max_tokens=3000,
        temperature=0,
        stop=["\n\n"]
    )

    output = response.choices[0].message.content.strip()
    # Use the original filename to create the output JSON file
    output_filename = f"extracted_{Path(original_filename).stem}.json"
    output_file_path = Path("output") / output_filename
    output_file_path.parent.mkdir(exist_ok=True)  # Create output directory if it doesn't exist
    
    with open(output_file_path, "w", encoding="utf-8") as f:
        f.write(output)
    print(f"Le contenu a été sauvegardé dans '{output_file_path}'.")

def process_files(input_dir="Docs"):
    input_dir = Path(input_dir)
    if not input_dir.exists():
        raise FileNotFoundError(f"Directory {input_dir} does not exist")
    
    # Create necessary directories
    temp_images_dir = Path('temp_images')
    temp_images_dir.mkdir(exist_ok=True)
    
    ocr_api_key = os.getenv('OCR_API_KEY')
    together_api_key = os.getenv('TOGETHER_API_KEY')

    if not ocr_api_key or not together_api_key:
        raise ValueError("API keys are not set. Please set the OCR_API_KEY and TOGETHER_API_KEY environment variables.")

    # Process all files in the input directory
    for file_path in input_dir.iterdir():
        if file_path.is_file():
            ext = file_path.suffix.lower()
            print(f"Processing file: {file_path.name}")
            
            if ext in ['.pdf', '.png', '.jpg', '.jpeg']:
                # Create a unique output file for each document
                output_file = f"ocr_output_{file_path.stem}.md"
                ocr_output_file = run_llama_ocr(str(file_path), output_file, str(temp_images_dir), ocr_api_key)
                if ocr_output_file:
                    extract_info_from_ocr(ocr_output_file, together_api_key, file_path.name)
            else:
                # Process non-image/PDF files directly
                output_path = Path("output") / f"{file_path.stem}.json"
                output_path.parent.mkdir(exist_ok=True)
                convert_to_json(file_path, output_path)

if __name__ == "__main__":
    process_files()
